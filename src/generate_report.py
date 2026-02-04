import requests
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from config import TOKEN, BASE_URL, PAGE_LIMIT, BIN_COMPANY, FIN_YEAR, CONTRACT_STATUSES, CONTRACT_TYPES, TERMINATED_STATUSES, DATE_FROM, DATE_TO, REPORTS_DIR, QUARTER

def get_quarter_dates(year, quarter):
    \"\"\"Получение дат начала и конца квартала\"\"\"
    if not quarter:
        return f\"{year}-01-01\", f\"{year}-12-31\"
    
    quarter_ranges = {
        1: (f\"{year}-01-01\", f\"{year}-03-31\"),
        2: (f\"{year}-04-01\", f\"{year}-06-30\"),
        3: (f\"{year}-07-01\", f\"{year}-09-30\"),
        4: (f\"{year}-10-01\", f\"{year}-12-31\")
    }
    return quarter_ranges.get(quarter, (f\"{year}-01-01\", f\"{year}-12-31\"))

def get_contracts_for_report():
    \"\"\"Получение договоров для отчёта\"\"\"
    
    headers = {
        \"Authorization\": f\"Bearer {TOKEN}\",
        \"Content-Type\": \"application/json\"
    }
    
    all_contracts = []
    after = 0
    
    while True:
        query = {
            \"query\": \"\"\"
                query($limit: Int, $after: Int, $filter: ContractFiltersInput!) {
                    Contract(limit: $limit, after: $after, filter: $filter) {
                        id
                        contractNumber
                        contractSum
                        contractSumWnds
                        faktSum
                        finYear
                        signDate
                        refContractStatusId
                        refContractTypeId
                        Supplier {
                            nameRu
                        }
                        RefContractStatus {
                            nameRu
                        }
                        RefSubjectType {
                            nameRu
                        }
                        FaktTradeMethods {
                            nameRu
                        }
                        ContractUnits {
                            Plans {
                                amount
                            }
                        }
                    }
                }
            \"\"\",
            \"variables\": {
                \"limit\": PAGE_LIMIT,
                \"after\": after,
                \"filter\": {
                    \"customerBin\": BIN_COMPANY,
                    \"finYear\": FIN_YEAR,
                    \"refContractStatusId\": CONTRACT_STATUSES
                }
            }
        }
        
        response = requests.post(f\"{BASE_URL}/v3/graphql\", json=query, headers=headers)
        data = response.json()
        
        if \"errors\" in data:
            print(f\"Ошибка API: {data['errors']}\")
            break
        
        contracts = data.get(\"data\", {}).get(\"Contract\", [])
        if not contracts:
            break
            
        all_contracts.extend(contracts)
        
        page_info = data.get(\"extensions\", {}).get(\"pageInfo\", {})
        if not page_info.get(\"hasNextPage\", False):
            break
        after = page_info.get(\"lastId\", 0)
        
        print(f\"Загружено: {len(all_contracts)} договоров...\")
    
    return all_contracts

def get_terminated_contracts_count(quarter=None):
    \"\"\"Получение количества расторгнутых договоров (с фильтрацией по кварталу)\"\"\"

    headers = {
        \"Authorization\": f\"Bearer {TOKEN}\",
        \"Content-Type\": \"application/json\"
    }

    all_terminated = []
    after = 0

    while True:
        query = {
            \"query\": \"\"\"
                query($limit: Int, $after: Int, $filter: ContractFiltersInput!) {
                    Contract(limit: $limit, after: $after, filter: $filter) {
                        id
                        signDate
                    }
                }
            \"\"\",
            \"variables\": {
                \"limit\": PAGE_LIMIT,
                \"after\": after,
                \"filter\": {
                    \"customerBin\": BIN_COMPANY,
                    \"finYear\": FIN_YEAR,
                    \"refContractStatusId\": TERMINATED_STATUSES
                }
            }
        }

        response = requests.post(f\"{BASE_URL}/v3/graphql\", json=query, headers=headers)
        data = response.json()

        if \"errors\" in data:
            print(f\"Ошибка API: {data['errors']}\")
            break

        contracts = data.get(\"data\", {}).get(\"Contract\", [])
        if not contracts:
            break
            
        all_terminated.extend(contracts)

        page_info = data.get(\"extensions\", {}).get(\"pageInfo\", {})
        if not page_info.get(\"hasNextPage\", False):
            break
        after = page_info.get(\"lastId\", 0)

    # Фильтруем по кварталу
    if quarter:
        all_terminated = filter_by_quarter(all_terminated, quarter)
    
    return len(all_terminated)

def get_announcements_by_method(date_from, date_to):
    \"\"\"Получение объявлений и группировка по способам закупки\"\"\"

    headers = {
        \"Authorization\": f\"Bearer {TOKEN}\",
        \"Content-Type\": \"application/json\"
    }

    methods_count = defaultdict(int)
    after = 0

    while True:
        query = {
            \"query\": \"\"\"
                query($limit: Int, $after: Int, $filter: TrdBuyFiltersInput!) {
                    TrdBuy(limit: $limit, after: $after, filter: $filter) {
                        id
                        RefTradeMethods {
                            nameRu
                        }
                    }
                }
            \"\"\",
            \"variables\": {
                \"limit\": PAGE_LIMIT,
                \"after\": after,
                \"filter\": {
                    \"orgBin\": BIN_COMPANY,
                    \"publishDate\": [date_from, date_to]
                }
            }
        }

        response = requests.post(f\"{BASE_URL}/v3/graphql\", json=query, headers=headers)
        data = response.json()

        if \"errors\" in data:
            print(f\"Ошибка API (объявления): {data['errors']}\")
            break

        announcements = data.get(\"data\", {}).get(\"TrdBuy\", [])
        if not announcements:
            break

        for a in announcements:
            method = a.get(\"RefTradeMethods\", {}).get(\"nameRu\") if a.get(\"RefTradeMethods\") else \"Не указан\"
            methods_count[method] += 1

        page_info = data.get(\"extensions\", {}).get(\"pageInfo\", {})
        if not page_info.get(\"hasNextPage\", False):
            break
        after = page_info.get(\"lastId\", 0)

    return dict(methods_count)

def get_plan_amount(contract):
    \"\"\"Получение плановой суммы из пунктов плана\"\"\"
    units = contract.get(\"ContractUnits\", [])
    if not units:
        return 0
    total = 0
    for unit in units:
        plans = unit.get(\"Plans\")
        if plans and plans.get(\"amount\"):
            total += plans.get(\"amount\", 0)
    return total

def filter_by_quarter(contracts, quarter):
    \"\"\"Фильтрация договоров по кварталу (по дате подписания)\"\"\"
    if not quarter:
        return contracts
    
    quarter_months = {
        1: (1, 2, 3),
        2: (4, 5, 6),
        3: (7, 8, 9),
        4: (10, 11, 12)
    }
    months = quarter_ranges.get(quarter, ())
    
    filtered = []
    for c in contracts:
        sign_date = c.get(\"signDate\")
        if sign_date:
            month = int(sign_date[5:7])  # \"2024-03-15\" -> 3
            if month in months:
                filtered.append(c)
    return filtered

def aggregate_data(contracts):
    \"\"\"Агрегация данных по способам закупки и видам предмета\"\"\"

    # Таблица 1: по способам закупки
    methods_data = defaultdict(lambda: {\"plan_sum\": 0, \"contract_sum\": 0, \"actual_sum\": 0, \"count\": 0})

    # Таблица 2: по способам и видам
    methods_types_data = defaultdict(lambda: defaultdict(lambda: {\"count\": 0, \"sum\": 0}))

    # Итоги по видам
    types_data = defaultdict(lambda: 0)

    for c in contracts:
        method = c.get(\"FaktTradeMethods\", {}).get(\"nameRu\") if c.get(\"FaktTradeMethods\") else \"Не указан\"
        subject_type = c.get(\"RefSubjectType\", {}).get(\"nameRu\") if c.get(\"RefSubjectType\") else \"Не указан\"

        contract_sum = float(c.get(\"contractSum\", 0) or 0)
        fakt_sum = float(c.get(\"faktSum\", 0) or 0)
        plan_sum = float(get_plan_amount(c) or 0)

        # Для экономии: если есть фактическая сумма - используем её, иначе сумму договора
        actual_sum = fakt_sum if fakt_sum > 0 else contract_sum

        # Таблица 1
        methods_data[method][\"plan_sum\"] += plan_sum
        methods_data[method][\"contract_sum\"] += contract_sum
        methods_data[method][\"actual_sum\"] += actual_sum
        methods_data[method][\"count\"] += 1

        # Таблица 2
        methods_types_data[method][subject_type][\"count\"] += 1
        methods_types_data[method][subject_type][\"sum\"] += contract_sum

        # Итоги по видам
        types_data[subject_type] += contract_sum

    return methods_data, methods_types_data, types_data

def format_number(value):
    \"\"\"Форматирование числа в тыс. тенге (округление без знаков после запятой)\"\"\"
    return round(value / 1000)

def create_report(contracts, filename, terminated_count=0, announcements_data=None, ann_dates=None):
    \"\"\"Создание Excel-отчёта\"\"\"
    
    methods_data, methods_types_data, types_data = aggregate_data(contracts)

    total_contract_sum = sum(m[\"contract_sum\"] for m in methods_data.values())
    total_actual_sum = sum(m[\"actual_sum\"] for m in methods_data.values())
    total_plan_sum = sum(m[\"plan_sum\"] for m in methods_data.values())
    total_economy = total_plan_sum - total_actual_sum
    total_count = sum(m[\"count\"] for m in methods_data.values())
    
    wb = Workbook()
    ws = wb.active
    ws.title = \"Итоги закупок\"
    
    # Стили
    font_title = Font(name='Times New Roman', size=16, bold=True)
    font_bold = Font(name='Times New Roman', size=12, bold=True)
    font_normal = Font(name='Times New Roman', size=12)
    font_header = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    row = 1

    # Заголовок отчёта
    if QUARTER:
        report_title = f\"ИТОГИ ГОСУДАРСТВЕННЫХ ЗАКУПОК ЗА {QUARTER} КВАРТАЛ {FIN_YEAR} ГОДА\"
    else:
        report_title = f\"ИТОГИ ГОСУДАРСТВЕННЫХ ЗАКУПОК ЗА {FIN_YEAR} ГОД\"
    ws.cell(row=row, column=2, value=report_title)
    ws.cell(row=row, column=2).font = font_title
    ws.cell(row=row, column=2).alignment = alignment_center
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 25
    row += 2

    # Подзаголовок со статусами
    ws.cell(row=row, column=2, value=\"Статусы договоров: Исполнен, Частично исполнен, Действует\")
    ws.cell(row=row, column=2).font = font_normal
    ws.cell(row=row, column=2).alignment = alignment_center
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 2

    # Сводка
    ws.cell(row=row, column=2, value=f\"Фактическая сумма по итогам государственных закупок составляет {format_number(total_actual_sum):,} тыс. тенге (без НДС). Экономия составила {format_number(total_economy):,} тыс. тенге.\".replace(\",\", \" \"))
    ws.cell(row=row, column=2).font = font_bold
    ws.cell(row=row, column=2).alignment = alignment_left
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 30
    row += 2

    # Вид предмета закупок
    ws.cell(row=row, column=2, value=\"Вид предмета закупок (по закупкам не превышающие финансовый год):\")
    ws.cell(row=row, column=2).font = font_bold
    row += 1

    for subject_type, sum_val in types_data.items():
        ws.cell(row=row, column=2, value=f\"{subject_type} - {format_number(sum_val):,} тыс. тенге\".replace(\",\", \" \"))
        ws.cell(row=row, column=2).font = font_normal
        row += 1

    ws.cell(row=row, column=2, value=f\"ИТОГО - {format_number(total_contract_sum):,} тыс. тенге\".replace(\",\", \" \"))
    ws.cell(row=row, column=2).font = font_bold
    row += 2

    ws.cell(row=row, column=2, value=f\"Согласно видам по закупкам (по закупкам не превышающие финансовый год): заключено {total_count} договоров на общую сумму {format_number(total_contract_sum):,} тыс. тенге (статус договоров: исполнен/частично исполнен + действует).\".replace(\",\", \" \"))
    ws.cell(row=row, column=2).font = font_normal
    ws.cell(row=row, column=2).alignment = alignment_left
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 30
    row += 3

    # Таблица 1: по способам закупки
    ws.cell(row=row, column=2, value=\"Закупки по видам не превышающие финансовый год (тыс. тенге)\")
    ws.cell(row=row, column=2).font = font_bold
    row += 2

    # Заголовки таблицы 1
    headers1 = [\"№\", \"Способ закупок\", \"Планируемая сумма без НДС\", \"Планируемая сумма с НДС\", \"Фактическая сумма без НДС\", \"Фактическая сумма с НДС\", \"Экономия без НДС\", \"Экономия с НДС\"]
    for col, header in enumerate(headers1, 2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border
        cell.fill = header_fill
    ws.row_dimensions[row].height = 30
    row += 1

    # Данные таблицы 1
    NDS_RATE = 1.12
    for idx, (method, data) in enumerate(methods_data.items(), 1):
        economy = data[\"plan_sum\"] - data[\"actual_sum\"]
        plan_sum_val = format_number(data[\"plan_sum\"]) if data[\"plan_sum\"] > 0 else \"-\"
        plan_sum_nds = format_number(data[\"plan_sum\"] * NDS_RATE) if data[\"plan_sum\"] > 0 else \"-\"
        actual_sum_val = format_number(data[\"actual_sum\"])
        actual_sum_nds = format_number(data[\"actual_sum\"] * NDS_RATE)
        economy_val = format_number(economy) if economy != 0 else 0
        economy_nds = format_number(economy * NDS_RATE) if economy != 0 else 0
        values = [idx, method, plan_sum_val, plan_sum_nds, actual_sum_val, actual_sum_nds, economy_val, economy_nds]
        for col, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font_normal
            cell.border = thin_border
            if col == 2:
                cell.alignment = alignment_center
            elif col == 3:
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_right
            if col in [4, 5, 6, 7, 8, 9] and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        row += 1

    # Итого таблицы 1
    totals = [\"\", \"ИТОГО:\", format_number(total_plan_sum), format_number(total_plan_sum * NDS_RATE), format_number(total_actual_sum), format_number(total_actual_sum * NDS_RATE), format_number(total_economy), format_number(total_economy * NDS_RATE)]
    for col, val in enumerate(totals, 2):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font_bold
        cell.border = thin_border
        cell.fill = total_fill
        if col == 3:
            cell.alignment = alignment_left
        else:
            cell.alignment = alignment_right
        if col in [4, 5, 6, 7, 8, 9] and isinstance(val, (int, float)):
            cell.number_format = '#,##0'
    row += 1

    # Примечание о расчёте НДС
    ws.cell(row=row, column=2, value=\"* Суммы с НДС рассчитаны с применением ставки 12% для всех договоров, независимо от статуса плательщика НДС поставщика.\")
    ws.cell(row=row, column=2).font = Font(name='Times New Roman', size=10, italic=True)
    ws.cell(row=row, column=2).alignment = alignment_left
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    row += 1

    # Примечание о фактических суммах
    ws.cell(row=row, column=2, value=\"** Фактическая сумма без НДС: для плательщиков НДС указана сумма без НДС, для неплательщиков НДС — без изменений.\")
    ws.cell(row=row, column=2).font = Font(name='Times New Roman', size=10, italic=True)
    ws.cell(row=row, column=2).alignment = alignment_left
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    row += 3

    # Таблица 2: по способам и видам
    headers2 = [\"№\", \"Способ закупки/вид закупки\", \"Количество договоров\", \"Общая сумма договоров без НДС\"]
    for col, header in enumerate(headers2, 2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border
        cell.fill = header_fill
    ws.row_dimensions[row].height = 30
    row += 1

    row_num = 1
    for method, types in methods_types_data.items():
        # Строка способа закупки
        cell = ws.cell(row=row, column=2, value=row_num)
        cell.font = font_bold
        cell.alignment = alignment_center
        cell.border = thin_border
        cell = ws.cell(row=row, column=3, value=method)
        cell.font = font_bold
        cell.alignment = alignment_left
        cell.border = thin_border
        for col in [4, 5]:
            cell = ws.cell(row=row, column=col, value=\"\")
            cell.border = thin_border
        row += 1
        row_num += 1

        # Строки видов предмета
        for subject_type, data in types.items():
            cell = ws.cell(row=row, column=2, value=\"\")
            cell.border = thin_border
            cell = ws.cell(row=row, column=3, value=subject_type)
            cell.font = font_normal
            cell.alignment = alignment_left
            cell.border = thin_border
            cell = ws.cell(row=row, column=4, value=data[\"count\"])
            cell.font = font_normal
            cell.alignment = alignment_right
            cell.border = thin_border
            cell = ws.cell(row=row, column=5, value=format_number(data[\"sum\"]))
            cell.font = font_normal
            cell.alignment = alignment_right
            cell.border = thin_border
            cell.number_format = '#,##0'
            row += 1

    # Итого таблицы 2
    cell = ws.cell(row=row, column=2, value=\"\")
    cell.border = thin_border
    cell.fill = total_fill
    cell = ws.cell(row=row, column=3, value=\"ИТОГО\")
    cell.font = font_bold
    cell.alignment = alignment_left
    cell.border = thin_border
    cell.fill = total_fill
    cell = ws.cell(row=row, column=4, value=total_count)
    cell.font = font_bold
    cell.alignment = alignment_right
    cell.border = thin_border
    cell.fill = total_fill
    cell = ws.cell(row=row, column=5, value=format_number(total_contract_sum))
    cell.font = font_bold
    cell.alignment = alignment_right
    cell.border = thin_border
    cell.fill = total_fill
    cell.number_format = '#,##0'
    row += 3

    # Информация о расторгнутых договорах
    ws.cell(row=row, column=2, value=f\"Количество расторгнутых договоров: {terminated_count}\")
    ws.cell(row=row, column=2).font = font_bold
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 2

    # Примечание о расчёте экономии
    ws.cell(row=row, column=2, value=\"Примечание: Экономия = Плановая сумма - Фактическая сумма (если факт > 0, иначе сумма договора)\")
    ws.cell(row=row, column=2).font = Font(name='Times New Roman', size=10, italic=True)
    ws.cell(row=row, column=2).alignment = alignment_left
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 30
    row += 2

    # Сводка по объявлениям
    if announcements_data:
        ann_period = f\"{ann_dates[0]} - {ann_dates[1]}\" if ann_dates else \"н/д\"
        ws.cell(row=row, column=2, value=f\"ОБЪЯВЛЕНИЯ О ЗАКУПКАХ за период {ann_period}\")
        ws.cell(row=row, column=2).font = font_bold
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 2

        # Заголовки
        ann_headers = [\"№\", \"Способ закупки\", \"Количество\"]
        for col, header in enumerate(ann_headers, 2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = font_header
            cell.alignment = alignment_center
            cell.border = thin_border
            cell.fill = header_fill
        ws.row_dimensions[row].height = 25
        row += 1

        # Данные
        total_ann = 0
        for idx, (method, count) in enumerate(sorted(announcements_data.items(), key=lambda x: -x[1]), 1):
            values = [idx, method, count]
            for col, val in enumerate(values, 2):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = font_normal
                cell.border = thin_border
                if col == 2:
                    cell.alignment = alignment_center
                elif col == 3:
                    cell.alignment = alignment_left
                else:
                    cell.alignment = alignment_right
            total_ann += count
            row += 1

        # Итого
        totals_ann = [\"\", \"ИТОГО\", total_ann]
        for col, val in enumerate(totals_ann, 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font_bold
            cell.border = thin_border
            cell.fill = total_fill
            if col == 3:
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_right

    # Автоширина колонок
    ws.column_dimensions['A'].width = 2.5
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    # Настройки печати: альбомная ориентация, умещение по ширине
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Не ограничивать по высоте
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    wb.save(filename)
    print(f\"Отчёт сохранён: {filename}\")

if __name__ == \"__main__\":
    print(f\"Генерация отчёта за {FIN_YEAR} год для заказчика {BIN_COMPANY}...\")
    print(f\"Фильтр: статусы {CONTRACT_STATUSES}, типы договоров {CONTRACT_TYPES}\")

    contracts = get_contracts_for_report()
    terminated_count = get_terminated_contracts_count(QUARTER)
    print(f\"Расторгнутых договоров: {terminated_count}\")

    # Определяем период для объявлений
    if QUARTER:
        ann_date_from, ann_date_to = get_quarter_dates(FIN_YEAR, QUARTER)
    else:
        ann_date_from, ann_date_to = get_quarter_dates(FIN_YEAR, None)  # Весь год
    print(f\"Загрузка объявлений за период {ann_date_from} - {ann_date_to}...\")
    announcements_data = get_announcements_by_method(ann_date_from, ann_date_to)
    print(f\"Объявлений: {sum(announcements_data.values())}\")

    if contracts:
        # Фильтрация по кварталу
        if QUARTER:
            contracts = filter_by_quarter(contracts, QUARTER)
            print(f\"После фильтрации по {QUARTER} кварталу: {len(contracts)} договоров\")
            filename = os.path.join(REPORTS_DIR, f\"report_{FIN_YEAR}_Q{QUARTER}.xlsx\")
        else:
            filename = os.path.join(REPORTS_DIR, f\"report_{FIN_YEAR}.xlsx\")
        create_report(contracts, filename, terminated_count, announcements_data, (ann_date_from, ann_date_to))
        print(f\"\\nГотово! Найдено договоров: {len(contracts)}\")
    else:
        print(\"Договоры не найдены.\")
