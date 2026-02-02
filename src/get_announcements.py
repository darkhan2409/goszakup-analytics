import requests
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from config import TOKEN, BASE_URL, PAGE_LIMIT, BIN_COMPANY, DATE_FROM, DATE_TO, REPORTS_DIR

def get_announcements(date_from, date_to):
    """Получение объявлений о закупках через GraphQL за период"""

    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json"
    }

    all_announcements = []
    after = 0

    while True:
        query = {
            "query": """
                query($limit: Int, $after: Int, $filter: TrdBuyFiltersInput!) {
                    TrdBuy(limit: $limit, after: $after, filter: $filter) {
                        id
                        RefTradeMethods {
                            nameRu
                        }
                    }
                }
            """,
            "variables": {
                "limit": PAGE_LIMIT,
                "after": after,
                "filter": {
                    "orgBin": BIN_COMPANY,
                    "publishDate": [date_from, date_to]
                }
            }
        }

        response = requests.post(f"{BASE_URL}/v3/graphql", json=query, headers=headers)
        data = response.json()

        if "errors" in data:
            print(f"Ошибка API: {data['errors']}")
            break

        announcements = data.get("data", {}).get("TrdBuy", [])
        if not announcements:
            break

        all_announcements.extend(announcements)

        page_info = data.get("extensions", {}).get("pageInfo", {})
        if not page_info.get("hasNextPage", False):
            break
        after = page_info.get("lastId", 0)

    return all_announcements

def count_by_method(announcements):
    """Подсчёт объявлений по способам закупки"""
    methods_count = defaultdict(int)

    for a in announcements:
        method = a.get("RefTradeMethods", {}).get("nameRu") if a.get("RefTradeMethods") else "Не указан"
        methods_count[method] += 1

    return methods_count

def save_to_excel(methods_count, total_count, filename):
    """Сохранение отчета по объявлениям в Excel с форматированием"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Объявления о закупках"
    
    # Стили
    font_title = Font(name='Times New Roman', size=14, bold=True)
    font_bold = Font(name='Times New Roman', size=12, bold=True)
    font_normal = Font(name='Times New Roman', size=12)
    font_header = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_right = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    row = 1
    
    # Заголовок
    ws.cell(row=row, column=2, value=f"ОБЪЯВЛЕНИЯ О ЗАКУПКАХ")
    ws.cell(row=row, column=2).font = font_title
    ws.cell(row=row, column=2).alignment = alignment_center
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Период и заказчик
    ws.cell(row=row, column=2, value=f"Период: {DATE_FROM} - {DATE_TO}")
    ws.cell(row=row, column=2).font = font_normal
    ws.cell(row=row, column=2).alignment = alignment_center
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 1
    
    ws.cell(row=row, column=2, value=f"БИН заказчика: {BIN_COMPANY}")
    ws.cell(row=row, column=2).font = font_normal
    ws.cell(row=row, column=2).alignment = alignment_center
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 2
    
    # Заголовки таблицы
    headers = ["№", "Способ закупки", "Количество"]
    for col, header in enumerate(headers, 2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border
        cell.fill = header_fill
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Данные (сортировка по убыванию количества)
    for idx, (method, count) in enumerate(sorted(methods_count.items(), key=lambda x: -x[1]), 1):
        values = [idx, method, count]
        for col, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font_normal
            cell.border = thin_border
            if col == 2:
                cell.alignment = alignment_center
            elif col == 3:
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_right
        row += 1
    
    # Итого
    totals = ["", "ИТОГО", total_count]
    for col, val in enumerate(totals, 2):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font_bold
        cell.border = thin_border
        cell.fill = total_fill
        if col == 3:
            cell.alignment = alignment_left
        else:
            cell.alignment = alignment_right
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 2.5
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 15
    
    wb.save(filename)
    print(f"Отчёт сохранён: {filename}")

if __name__ == "__main__":
    print(f"Период: {DATE_FROM} - {DATE_TO}")
    print(f"Заказчик: {BIN_COMPANY}\n")

    announcements = get_announcements(DATE_FROM, DATE_TO)

    if announcements:
        methods_count = count_by_method(announcements)

        print("Способ закупки                                    | Кол-во")
        print("-" * 60)
        for method, count in sorted(methods_count.items(), key=lambda x: -x[1]):
            print(f"{method[:48]:<48} | {count}")
        print("-" * 60)
        print(f"{'ИТОГО':<48} | {len(announcements)}")
        
        # Сохранение в Excel
        period_start = DATE_FROM.replace('-', '')
        period_end = DATE_TO.replace('-', '')
        filename = os.path.join(REPORTS_DIR, f"announcements_{BIN_COMPANY}_{period_start}_{period_end}.xlsx")
        save_to_excel(methods_count, len(announcements), filename)
        print(f"\nВсего объявлений: {len(announcements)}")
    else:
        print("Объявления не найдены.")
